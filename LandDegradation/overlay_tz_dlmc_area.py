"""
Overlay soil-type polygons with DLTB land-use polygons and summarize area in mu.

Target land classes are the second-level DLBM classes for cultivated land,
garden land, forest land, grassland, and other land.

Command line example:
  python LandDegradation/overlay_tz_dlmc_area.py
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
MU_DIVISOR = 666.6666667


TARGET_LAND_CLASSES = {
    "0101": ("01", "耕地", "水田"),
    "0102": ("01", "耕地", "水浇地"),
    "0103": ("01", "耕地", "旱地"),
    "0201": ("02", "园地", "果园"),
    "0202": ("02", "园地", "茶园"),
    "0203": ("02", "园地", "橡胶园"),
    "0204": ("02", "园地", "其他园地"),
    "0301": ("03", "林地", "乔木林地"),
    "0302": ("03", "林地", "竹林地"),
    "0305": ("03", "林地", "灌木林地"),
    "0307": ("03", "林地", "其他林地"),
    "0401": ("04", "草地", "天然牧草地"),
    "0403": ("04", "草地", "人工牧草地"),
    "0404": ("04", "草地", "其他草地"),
    "1202": ("12", "其他土地", "设施农用地"),
    "1204": ("12", "其他土地", "盐碱地"),
    "1205": ("12", "其他土地", "沙地"),
    "1206": ("12", "其他土地", "裸土地"),
}


MATRIX_CLASS_ORDER = [
    "0101",
    "0102",
    "0103",
    "0201",
    "0204",
    "0301",
    "0307",
    "0404",
    "1202",
    "1205",
    "1206",
]


BARRIER_TYPE_BY_TZ = {
    "砂壤底黏石灰性潮土": "黏质型障碍",
    "砂壤夹黏石灰性潮土": "黏质型障碍",
    "砂壤体黏石灰性潮土": "黏质型障碍",
    "壤体黏石灰性潮土": "黏质型障碍",
    "壤底黏石灰性潮土": "黏质型障碍",
    "壤夹黏石灰性潮土": "黏质型障碍",
    "壤底黏脱潮土": "黏质型障碍",
    "壤底黏泥砂质潮褐土": "黏质型障碍",
    "壤体黏泥砂质潮褐土": "黏质型障碍",
    "砂壤底黏泥砂质潮褐土": "黏质型障碍",
    "砂壤体黏泥砂质潮褐土": "黏质型障碍",
    "壤底砂石灰性潮土": "夹砂型障碍",
    "壤体砂石灰性潮土": "夹砂型障碍",
    "黏壤夹砂石灰性潮土": "夹砂型障碍",
    "壤底砂泥A砂质潮褐土": "夹砂型障碍",
    "壤底砂泥 A砂质潮褐土": "夹砂型障碍",
    "壤底黏底砂姜石灰性潮土": "砂姜黏质复合障碍",
    "壤体黏底砂姜石灰性潮土": "砂姜黏质复合障碍",
}


def require_dependencies() -> None:
    missing = []
    for package, import_name in [
        ("geopandas", "geopandas"),
        ("shapely", "shapely"),
        ("pandas", "pandas"),
        ("pyogrio", "pyogrio"),
    ]:
        try:
            __import__(import_name)
        except ImportError:
            missing.append(package)

    if missing:
        print("缺少必要 Python 包：", ", ".join(missing), file=sys.stderr)
        print(
            "请先安装依赖：python -m pip install -r LandDegradation/requirements.txt",
            file=sys.stderr,
        )
        raise SystemExit(1)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="统计每个土种 TZ 内目标二级地类的面积，单位为亩。")
    parser.add_argument(
        "--soil",
        default=str(PROJECT_DIR / "Data" / "安国市三普土壤图.shp"),
        help="三普土壤图 shp 路径，需包含 TZ 字段。",
    )
    parser.add_argument(
        "--dltb",
        default=str(PROJECT_DIR / "Data" / "安国DLTB.shp"),
        help="DLTB shp 路径，需包含 DLBM 字段。",
    )
    parser.add_argument(
        "--out-dir",
        default=str(SCRIPT_DIR / "Output"),
        help="输出目录。默认输出到 LandDegradation/Output。",
    )
    parser.add_argument(
        "--area-crs",
        default=None,
        help=(
            "用于面积计算的目标坐标系，例如 EPSG:4547。"
            "默认使用三普土壤图自身坐标系；若三普土壤图为经纬度，则改用 DLTB 坐标系。"
        ),
    )
    parser.add_argument(
        "--save-intersection",
        action="store_true",
        help="同时保存叠加后的面要素 GeoPackage，便于 GIS 软件检查。",
    )
    return parser.parse_args()


def read_layer(path: Path, columns: list[str]):
    import geopandas as gpd

    if not path.exists():
        raise FileNotFoundError(f"找不到文件：{path}")

    return gpd.read_file(path, columns=columns, encoding="UTF-8", engine="pyogrio")


def validate_layer(gdf, path: Path, required_fields: list[str], layer_name: str) -> None:
    if gdf.empty:
        raise ValueError(f"{layer_name} 没有任何要素：{path}")

    missing = [field for field in required_fields if field not in gdf.columns]
    if missing:
        fields = ", ".join([c for c in gdf.columns if c != "geometry"])
        raise ValueError(f"{layer_name} 缺少字段 {', '.join(missing)}。当前字段：{fields}")

    if gdf.crs is None:
        raise ValueError(f"{layer_name} 缺少坐标系信息，请先检查 .prj 文件：{path}")


def repair_geometry(gdf):
    try:
        from shapely.validation import make_valid
    except ImportError:
        make_valid = None

    gdf = gdf.copy()
    gdf = gdf[~gdf.geometry.isna() & ~gdf.geometry.is_empty].copy()
    invalid = ~gdf.geometry.is_valid
    if invalid.any():
        if make_valid is None:
            gdf.loc[invalid, "geometry"] = gdf.loc[invalid, "geometry"].buffer(0)
        else:
            gdf.loc[invalid, "geometry"] = gdf.loc[invalid, "geometry"].apply(make_valid)

    gdf = gdf.explode(index_parts=False, ignore_index=True)
    gdf = gdf[~gdf.geometry.isna() & ~gdf.geometry.is_empty].copy()
    return gdf


def choose_area_crs(soil, dltb, area_crs: str | None):
    if area_crs:
        return area_crs
    if soil.crs and not soil.crs.is_geographic:
        return soil.crs
    if dltb.crs and not dltb.crs.is_geographic:
        return dltb.crs
    raise ValueError("两个图层都是经纬度坐标系，不能直接算面积；请用 --area-crs 指定投影坐标系。")


def normalize_second_level_code(value) -> str:
    if value is None:
        return ""

    code = str(value).strip()
    if not code:
        return ""

    if code.endswith(".0"):
        code = code[:-2]

    if code.isdigit() and len(code) < 4:
        code = code.zfill(4)

    return code[:4]


def get_barrier_type(tz: str) -> str:
    return BARRIER_TYPE_BY_TZ.get(str(tz).strip(), "无")


def add_target_land_class_fields(dltb):
    dltb = dltb.copy()
    dltb["二级类编码"] = dltb["DLBM"].apply(normalize_second_level_code)
    dltb = dltb[dltb["二级类编码"].isin(TARGET_LAND_CLASSES)].copy()

    class_info = dltb["二级类编码"].map(TARGET_LAND_CLASSES)
    dltb["一级类编码"] = class_info.apply(lambda item: item[0])
    dltb["一级类名称"] = class_info.apply(lambda item: item[1])
    dltb["二级类名称"] = class_info.apply(lambda item: item[2])
    return dltb


def integerize_columns(df, columns: list[str]):
    df = df.copy()
    for column in columns:
        df[column] = df[column].round(0).astype("int64")
    return df


def safe_name_part(value: str | Path, fallback: str = "未命名") -> str:
    name = Path(value).stem if value else fallback
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", str(name)).strip(" ._")
    return name or fallback


def build_output_prefix(soil_path: str | Path, dltb_path: str | Path, suffix: str = "TZ_DLBM二级类_面积统计_亩_整数") -> str:
    soil_name = safe_name_part(soil_path, "土壤图")
    dltb_name = safe_name_part(dltb_path, "DLTB")
    return f"{soil_name}__{dltb_name}__{suffix}"


def build_matrix(summary, drop_zero_columns: bool = True):
    pivot = summary.pivot_table(
        index=["障碍类型", "TZ"],
        columns="二级类编码",
        values="面积_亩",
        aggfunc="sum",
        fill_value=0,
    )

    for code in MATRIX_CLASS_ORDER:
        if code not in pivot.columns:
            pivot[code] = 0

    class_order = MATRIX_CLASS_ORDER
    if drop_zero_columns:
        class_order = [code for code in MATRIX_CLASS_ORDER if int(round(pivot[code].sum())) != 0]

    pivot = pivot[class_order].round(0).astype("int64").reset_index()
    return pivot, class_order


def write_matrix_sheet(writer, summary) -> None:
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter

    sheet_name = "TZ_DLBM矩阵_亩"
    matrix, class_order = build_matrix(summary)
    start_col = 3
    total_col = start_col + len(class_order)

    workbook = writer.book
    worksheet = workbook.create_sheet(sheet_name)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)

    worksheet.cell(row=1, column=1, value="障碍类型")
    worksheet.cell(row=1, column=2, value="土种")
    worksheet.cell(row=1, column=start_col, value="土地利用类型")
    worksheet.cell(row=1, column=total_col, value="总计")
    worksheet.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=3, end_column=2)
    if class_order:
        worksheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=total_col - 1)
    worksheet.merge_cells(start_row=1, start_column=total_col, end_row=3, end_column=total_col)

    for offset, code in enumerate(class_order):
        col = start_col + offset
        _, first_name, second_name = TARGET_LAND_CLASSES[code]
        worksheet.cell(row=2, column=col, value=first_name)
        worksheet.cell(row=3, column=col, value=second_name)

    merge_start = start_col
    while merge_start < total_col:
        first_name = worksheet.cell(row=2, column=merge_start).value
        merge_end = merge_start
        while (
            merge_end + 1 < total_col
            and worksheet.cell(row=2, column=merge_end + 1).value == first_name
        ):
            merge_end += 1
        if merge_end > merge_start:
            worksheet.merge_cells(
                start_row=2,
                start_column=merge_start,
                end_row=2,
                end_column=merge_end,
            )
        merge_start = merge_end + 1

    for row_index, (_, row) in enumerate(matrix.iterrows(), start=4):
        worksheet.cell(row=row_index, column=1, value=row["障碍类型"])
        worksheet.cell(row=row_index, column=2, value=row["TZ"])
        row_total = 0
        for offset, code in enumerate(class_order):
            value = int(row[code])
            row_total += value
            worksheet.cell(row=row_index, column=start_col + offset, value=value)
        worksheet.cell(row=row_index, column=total_col, value=row_total)

    first_data_row = 4
    last_data_row = 3 + len(matrix)
    merge_start = first_data_row
    while merge_start <= last_data_row:
        barrier_type = worksheet.cell(row=merge_start, column=1).value
        merge_end = merge_start
        while (
            merge_end + 1 <= last_data_row
            and worksheet.cell(row=merge_end + 1, column=1).value == barrier_type
        ):
            merge_end += 1
        if merge_end > merge_start:
            worksheet.merge_cells(
                start_row=merge_start,
                start_column=1,
                end_row=merge_end,
                end_column=1,
            )
        merge_start = merge_end + 1

    worksheet.column_dimensions["A"].width = 16
    worksheet.column_dimensions["B"].width = 28
    for col in range(start_col, total_col + 1):
        worksheet.column_dimensions[get_column_letter(col)].width = 12

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = center
            cell.border = border
            if cell.row <= 3:
                cell.font = bold


def write_barrier_stats_sheet(writer, summary) -> None:
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter

    sheet_name = "障碍土种统计"
    matrix, class_order = build_matrix(summary, drop_zero_columns=False)
    matrix = matrix[matrix["障碍类型"] != "无"].copy()
    class_order = [code for code in class_order if not matrix.empty and int(matrix[code].sum()) != 0]
    matrix = matrix[["障碍类型", "TZ", *class_order]].copy()
    start_col = 3
    total_col = start_col + len(class_order)
    percent_col = total_col + 1

    workbook = writer.book
    worksheet = workbook.create_sheet(sheet_name)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)

    worksheet.cell(row=1, column=1, value="障碍类型")
    worksheet.cell(row=1, column=2, value="土种")
    worksheet.cell(row=1, column=start_col, value="土地利用类型")
    worksheet.cell(row=1, column=total_col, value="总计/亩")
    worksheet.cell(row=1, column=percent_col, value="占比/%")
    worksheet.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=3, end_column=2)
    if class_order:
        worksheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=total_col - 1)
    worksheet.merge_cells(start_row=1, start_column=total_col, end_row=3, end_column=total_col)
    worksheet.merge_cells(start_row=1, start_column=percent_col, end_row=3, end_column=percent_col)

    for offset, code in enumerate(class_order):
        col = start_col + offset
        _, first_name, second_name = TARGET_LAND_CLASSES[code]
        worksheet.cell(row=2, column=col, value=first_name)
        worksheet.cell(row=3, column=col, value=second_name)

    merge_start = start_col
    while merge_start < total_col:
        first_name = worksheet.cell(row=2, column=merge_start).value
        merge_end = merge_start
        while merge_end + 1 < total_col and worksheet.cell(row=2, column=merge_end + 1).value == first_name:
            merge_end += 1
        if merge_end > merge_start:
            worksheet.merge_cells(
                start_row=2,
                start_column=merge_start,
                end_row=2,
                end_column=merge_end,
            )
        merge_start = merge_end + 1

    if not matrix.empty:
        matrix["总计/亩"] = matrix[class_order].sum(axis=1).astype("int64")
        grand_total = int(matrix["总计/亩"].sum())
        if grand_total:
            matrix["占比/%"] = (matrix["总计/亩"] / grand_total * 100).round(2)
        else:
            matrix["占比/%"] = 0.0
    else:
        grand_total = 0
        matrix["总计/亩"] = []
        matrix["占比/%"] = []

    for row_index, (_, row) in enumerate(matrix.iterrows(), start=4):
        worksheet.cell(row=row_index, column=1, value=row["障碍类型"])
        worksheet.cell(row=row_index, column=2, value=row["TZ"])
        for offset, code in enumerate(class_order):
            worksheet.cell(row=row_index, column=start_col + offset, value=int(row[code]))
        worksheet.cell(row=row_index, column=total_col, value=int(row["总计/亩"]))
        worksheet.cell(row=row_index, column=percent_col, value=float(row["占比/%"]))

    first_data_row = 4
    last_data_row = 3 + len(matrix)
    merge_start = first_data_row
    while merge_start <= last_data_row:
        barrier_type = worksheet.cell(row=merge_start, column=1).value
        merge_end = merge_start
        while (
            merge_end + 1 <= last_data_row
            and worksheet.cell(row=merge_end + 1, column=1).value == barrier_type
        ):
            merge_end += 1
        if merge_end > merge_start:
            worksheet.merge_cells(
                start_row=merge_start,
                start_column=1,
                end_row=merge_end,
                end_column=1,
            )
        merge_start = merge_end + 1

    total_row = max(last_data_row + 1, 4)
    percent_row = total_row + 1
    class_totals = matrix[class_order].sum(axis=0).astype("int64") if not matrix.empty else {}

    worksheet.cell(row=total_row, column=1, value="总计/亩")
    worksheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    for offset, code in enumerate(class_order):
        value = int(class_totals[code]) if code in class_totals else 0
        worksheet.cell(row=total_row, column=start_col + offset, value=value)
    worksheet.cell(row=total_row, column=total_col, value=grand_total)
    worksheet.cell(row=total_row, column=percent_col, value=100 if grand_total else 0)

    worksheet.cell(row=percent_row, column=1, value="占比/%")
    worksheet.merge_cells(start_row=percent_row, start_column=1, end_row=percent_row, end_column=2)
    for offset, code in enumerate(class_order):
        value = int(class_totals[code]) if code in class_totals else 0
        percent = round(value / grand_total * 100, 2) if grand_total else 0
        worksheet.cell(row=percent_row, column=start_col + offset, value=percent)
    worksheet.cell(row=percent_row, column=total_col, value=100 if grand_total else 0)

    worksheet.column_dimensions["A"].width = 16
    worksheet.column_dimensions["B"].width = 28
    for col in range(start_col, percent_col + 1):
        worksheet.column_dimensions[get_column_letter(col)].width = 12

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = center
            cell.border = border
            if cell.row <= 3 or cell.row in (total_row, percent_row):
                cell.font = bold
            if cell.row >= 4 and cell.column == percent_col:
                cell.number_format = "0.00"
            if cell.row == percent_row and start_col <= cell.column <= total_col:
                cell.number_format = "0.00"


def summarize(soil, dltb):
    import geopandas as gpd

    soil = soil[["TZ", "geometry"]].copy()
    dltb = dltb[["DLBM", "geometry"]].copy()

    soil["TZ"] = soil["TZ"].fillna("未填写").astype(str).str.strip()
    soil["障碍类型"] = soil["TZ"].apply(get_barrier_type)
    dltb = add_target_land_class_fields(dltb)

    if dltb.empty:
        raise ValueError("DLTB 中没有匹配目标二级类编码的要素，请检查 DLBM 字段。")

    overlay = gpd.overlay(soil, dltb, how="intersection", keep_geom_type=True)
    overlay = overlay[~overlay.geometry.isna() & ~overlay.geometry.is_empty].copy()
    overlay["面积_亩"] = overlay.geometry.area / MU_DIVISOR
    overlay = overlay[overlay["面积_亩"] > 0].copy()

    group_fields = ["障碍类型", "TZ", "一级类编码", "一级类名称", "二级类编码", "二级类名称"]
    summary = (
        overlay.groupby(group_fields, dropna=False, as_index=False)["面积_亩"]
        .sum()
        .sort_values(["障碍类型", "TZ", "一级类编码", "二级类编码"], ascending=[True, True, True, True])
    )

    tz_total = summary.groupby(["障碍类型", "TZ"], as_index=False)["面积_亩"].sum()
    tz_total = tz_total.rename(columns={"面积_亩": "土种目标地类总面积_亩"})
    summary = summary.merge(tz_total, on=["障碍类型", "TZ"], how="left")
    summary["占土种目标地类面积比例_%"] = summary["面积_亩"] / summary["土种目标地类总面积_亩"] * 100.0

    summary = integerize_columns(
        summary,
        ["面积_亩", "土种目标地类总面积_亩", "占土种目标地类面积比例_%"],
    )
    overlay["面积_亩"] = overlay["面积_亩"].round(0).astype("int64")
    return summary, overlay


def _write_gpkg_layer(gdf, output_path: Path, layer_name: str) -> None:
    import pyogrio

    pyogrio.write_dataframe(
        gdf,
        output_path,
        layer=layer_name,
        driver="GPKG",
        layer_options={"SPATIAL_INDEX": "YES"},
    )


def build_arcgis_mapping_layers(overlay):
    mapping = overlay.copy()
    mapping = mapping.rename(
        columns={
            "一级类名称": "一级地类名称",
            "二级类名称": "二级地类名称",
            "TZ": "土种",
        }
    )
    mapping["障碍图斑"] = mapping["障碍类型"].apply(lambda value: "否" if value == "无" else "是")

    keep_cols = [
        "一级地类名称",
        "二级地类名称",
        "二级类编码",
        "障碍类型",
        "障碍图斑",
        "土种",
        "面积_亩",
        "geometry",
    ]
    mapping = mapping[keep_cols].copy()
    barrier = mapping[mapping["障碍图斑"] == "是"].copy()
    return mapping, barrier


def write_arcgis_mapping_database(overlay, output_path: Path) -> Path | None:
    mapping, barrier = build_arcgis_mapping_layers(overlay)
    if barrier.empty:
        print("没有提取到障碍类型不为“无”的图斑，已跳过 ArcGIS 制图数据库输出。")
        return None

    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists():
        output_path.unlink()

    _write_gpkg_layer(mapping, output_path, "制图结果_含障碍图斑")
    _write_gpkg_layer(barrier, output_path, "土体障碍图斑")

    print(f"已输出 ArcGIS 制图数据库 GeoPackage：{output_path}")
    return output_path


def write_outputs(
    summary,
    overlay,
    out_dir: Path,
    save_intersection: bool,
    prefix: str = "土种_DLBM二级类_面积统计_亩",
) -> dict[str, Path | None]:
    out_dir.mkdir(parents=True, exist_ok=True)

    csv_path = out_dir / f"{prefix}.csv"
    xlsx_path = out_dir / f"{prefix}.xlsx"
    gpkg_path = out_dir / f"{prefix}_叠加结果.gpkg"
    arcgis_gpkg_path = out_dir / f"{prefix}_ArcGIS制图数据库.gpkg"

    summary.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"已输出 CSV：{csv_path}")

    try:
        import pandas as pd

        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            summary.to_excel(writer, sheet_name="TZ_DLBM二级类面积_亩", index=False)
            write_matrix_sheet(writer, summary)
            write_barrier_stats_sheet(writer, summary)
        print(f"已输出 Excel：{xlsx_path}")
    except ImportError:
        xlsx_path = None
        print("未安装 openpyxl，已跳过 Excel 输出；CSV 已正常生成。")

    gpkg_output = None
    if save_intersection:
        keep_cols = [
            "障碍类型",
            "TZ",
            "DLBM",
            "二级类编码",
            "一级类编码",
            "一级类名称",
            "二级类名称",
            "面积_亩",
            "geometry",
        ]
        overlay[keep_cols].to_file(
            gpkg_path, layer="intersection", driver="GPKG", encoding="UTF-8"
        )
        gpkg_output = gpkg_path
        print(f"已输出叠加面 GeoPackage：{gpkg_path}")

    arcgis_gpkg_output = write_arcgis_mapping_database(overlay, arcgis_gpkg_path)

    return {
        "csv": csv_path,
        "xlsx": xlsx_path,
        "gpkg": gpkg_output,
        "arcgis_gpkg": arcgis_gpkg_output,
    }


def run_analysis(
    soil_path: str | Path,
    dltb_path: str | Path,
    out_dir: str | Path = SCRIPT_DIR / "Output",
    area_crs: str | None = None,
    save_intersection: bool = False,
    prefix: str | None = None,
) -> tuple[object, dict[str, Path | None]]:
    soil_path = Path(soil_path)
    dltb_path = Path(dltb_path)
    out_dir = Path(out_dir)
    if prefix is None:
        prefix = build_output_prefix(soil_path, dltb_path)

    soil = read_layer(soil_path, ["TZ", "geometry"])
    dltb = read_layer(dltb_path, ["DLBM", "geometry"])

    validate_layer(soil, soil_path, ["TZ"], "三普土壤图")
    validate_layer(dltb, dltb_path, ["DLBM"], "DLTB")

    target_crs = choose_area_crs(soil, dltb, area_crs)
    soil = repair_geometry(soil.to_crs(target_crs))
    dltb = repair_geometry(dltb.to_crs(target_crs))

    summary, overlay = summarize(soil, dltb)
    if summary.empty:
        raise ValueError("叠加结果为空：请检查两个图层是否在同一地理范围内。")

    outputs = write_outputs(summary, overlay, out_dir, save_intersection, prefix)
    return summary, outputs


def main() -> None:
    require_dependencies()
    args = parse_args()

    print("开始统计，面积单位：亩，结果不保留小数")
    run_analysis(
        soil_path=args.soil,
        dltb_path=args.dltb,
        out_dir=args.out_dir,
        area_crs=args.area_crs,
        save_intersection=args.save_intersection,
    )
    print("完成。")


if __name__ == "__main__":
    main()
